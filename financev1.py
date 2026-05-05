import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ====== תיק השקעות ======
PORTFOLIO = {
    "TDS": 0,
    "SPOK": 0,
    "VZ": 10,
    "SBGI": 0,
    "GTN": 15,
    "EVC": 0,
    "T": 13,

    "BKE": 0,
    "HVT": 0,
    "LEG": 0,
    "LZB": 0,
    "ETD": 16,
    "F": 25,

    "BGS": 0,
    "KHC": 0,
    "FLO": 0,
    "EPC": 0,
    "ADM": 5,
    "BG": 3,
    "PEP": 2,
    "MO": 5,

    "PBF": 0,
    "DK": 0,
    "CVI": 0,
    # "INT": 0,
    "SGU": 0,
    "CNX": 0,
    "COP": 2,
    "SU": 11,
    "XOM": 0,
    "CWEN-A": 13,
    "NEE": 3,

    "HOPE": 0,
    "NAVI": 0,
    "HWC": 0,
    "USB": 8,
    "BAC": 5,
    "FHI": 0,
    "SAR": 17,
    "TROW": 4,
    "FSK": 0,
    "ARCC": 0,
    "MAIN": 0,
    "GBDC": 0,

    "PINC": 0,
    "AVNS": 0,
    "MDT": 3,
    "PFE": 14,

    "LMT": 0,
    "RTX": 2,
    "AIR": 0,
    "ZIM": 28,
    "UPS": 5,
    "TRN": 10,
    "WERN": 0,
    "MMM": 1,
    "BRC": 0,

    "VSH": 0,
    "NSIT": 0,
    "SCSC": 0,
    "DGII": 0,
    "NTAP": 0,
    "GOOGL": 4,
    "MSFT": 2,
    "AAPL": 2,

    "CC": 0,
    "ARLP": 0,
    "RIO": 5,
    "HL": 0,
    "GEF": 0,
    "FMC": 11,
    "VALE": 29,
    "STLD": 1,
    "LYB": 2,

    "SBRA": 20,
    "OHI": 0,
    "PEB": 0,
    "SVC": 0,
    "AMH": 0,
    "APLE": 36,
    "SITC": 0,
    "O": 6,
    "AVB": 2,
    "GNL": 12,
    "VICI": 0,
    "CHMI": 19,
    "KIM": 0,
    "FRT": 0,
    "OLP": 15,
    "DEA": 16,
    "WSR": 0,

    "NWE": 6,
    "UTL": 0,
    "NJR": 7,
    "AVA": 10,
    "DUK": 2,
    "EIX": 6,
    "AWK": 0,
    "WTRG": 0,
    "CWT": 0,
    "MSEX": 0,
    "YORW": 0,
    "ARTNA": 0,
    "CWCO": 0,
    "GWRS": 0,

}

OUTPUT_FILE = "portfolio_tracker.xlsx"

SECTOR_COLORS = [
    "E8F0FE",  # blue
    "E6F4EA",  # green
    "FEF7E0",  # yellow
    "FCE8E6",  # red
    "F3E8FD",  # purple
    "E0F7FA",  # cyan
]


def fetch_stock(symbol, shares):
    t = yf.Ticker(symbol)
    info = t.info

    price = info.get("currentPrice")
    eps = info.get("trailingEps")
    pe = info.get("trailingPE")

    dividend = info.get("dividendRate")
    dividend_yield = info.get("dividendYield")
    if dividend_yield:
        dividend_yield = info.get("dividendYield")/100

    return {
        "Sector": info.get("sector", "Unknown"),
        "Symbol": symbol,
        "Company": info.get("shortName"),
        "Shares": shares,
        "Price": price,
        "Position Value": shares * price if price else None,
        "EPS": eps,
        "P/E": pe if pe else (price / eps if price and eps else None),
        "Industry": info.get("industry"),
        "Dividend": dividend,
        "Dividend Yield %": dividend_yield * 100 if dividend_yield else None,
        "Update Date": datetime.now().strftime("%Y-%m-%d"),
    }


def build_dataframe():
    rows = []
    for symbol, shares in PORTFOLIO.items():
        rows.append(fetch_stock(symbol, shares))

    df = pd.DataFrame(rows)
    df = df.sort_values(["Sector", "Symbol"])
    return df


def write_excel_with_sector_totals(df):
    final_rows = []
    for sector, group in df.groupby("Sector"):
        final_rows.extend(group.to_dict("records"))

        total_value = group["Position Value"].sum()
        final_rows.append(
            {
                "Sector": sector,
                "Symbol": f"TOTAL {sector}",
                "Company": "",
                "Shares": "",
                "Price": "",
                "Position Value": total_value,
                "P/E": "",
            }
        )

    final_df = pd.DataFrame(final_rows)
    final_df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    sector_color_map = {}
    color_idx = 0

    for row in range(2, ws.max_row + 1):
        sector = ws[f"A{row}"].value
        symbol = ws[f"B{row}"].value

        if sector not in sector_color_map:
            sector_color_map[sector] = SECTOR_COLORS[color_idx % len(SECTOR_COLORS)]
            color_idx += 1

        fill = PatternFill(
            start_color=sector_color_map[sector],
            end_color=sector_color_map[sector],
            fill_type="solid",
        )

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

        if isinstance(symbol, str) and symbol.startswith("TOTAL"):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)

    wb.save(OUTPUT_FILE)


def main():
    df = build_dataframe()
    write_excel_with_sector_totals(df)
    print(f"✅ Portfolio Excel created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
